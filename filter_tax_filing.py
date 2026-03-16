"""One-off: read tax_emails_raw.xlsx, keep only filing/official tax rows, write tax_emails_curated.xlsx."""
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

_BASE = Path(__file__).resolve().parent
raw_path = _BASE / "tax_emails_raw.xlsx"
out_path = _BASE / "tax_emails_curated.xlsx"

filing_keywords = (
    "property tax", "tax return", "irs", "1040", "w-2", "w2", "file taxes", "tax filing",
    "tax assessment", "department of revenue", "tax form", "tax authority", "income tax",
    "form 1040", "extension", "tax deadline", "filing deadline", "federal tax", "state tax",
    "tax refund", "1099", "tax year", "property tax assessment", "assessor", "tax bill",
)

wb = load_workbook(raw_path)
ws = wb.active
headers = [ws.cell(1, c).value for c in range(1, 9)]
rows = []
for r in range(2, ws.max_row + 1):
    row = {
        "provider": ws.cell(r, 1).value or "",
        "account": ws.cell(r, 2).value or "",
        "from": ws.cell(r, 3).value or "",
        "subject": ws.cell(r, 4).value or "",
        "date": ws.cell(r, 5).value or "",
        "snippet": ws.cell(r, 6).value or "",
        "link": ws.cell(r, 7).value or "",
        "query_used": ws.cell(r, 8).value or "",
    }
    rows.append(row)

kept = []
for r in rows:
    text = (r["subject"] + " " + r["snippet"] + " " + str(r["from"])).lower()
    if any(f in text for f in filing_keywords):
        kept.append(r)

out = Workbook()
sheet = out.active
sheet.title = "Emails"
headers_out = ["Provider", "Account", "From", "Subject", "Date", "Snippet", "Open link", "Search query used"]
for col, h in enumerate(headers_out, 1):
    sheet.cell(1, col, h)
    sheet.cell(1, col).font = Font(bold=True)
for i, r in enumerate(kept, 2):
    sheet.cell(i, 1, r["provider"])
    sheet.cell(i, 2, r["account"])
    sheet.cell(i, 3, r["from"])
    sheet.cell(i, 4, r["subject"])
    sheet.cell(i, 5, r["date"])
    sheet.cell(i, 6, r["snippet"])
    link = r["link"]
    cell = sheet.cell(i, 7, "Open in mail" if link else "")
    if link:
        cell.hyperlink = link
        cell.font = Font(color="0563C1", underline="single")
    sheet.cell(i, 8, r["query_used"])
for col in range(1, 9):
    sheet.column_dimensions[get_column_letter(col)].width = max(12, min(50, 25))
sheet.column_dimensions["F"].width = 50
out.save(out_path)
print(f"Filtered: {len(kept)} of {len(rows)} -> {out_path}")
