"""
Email MCP server: search Gmail or Outlook and export results to Excel (with links).
Gmail: gmail_credentials.json → gmail_token.json.
Outlook: OUTLOOK_CLIENT_ID → outlook_token_cache.bin. Optional second account: OUTLOOK_TOKEN_CACHE_2.
"""
from __future__ import annotations

import json
import os
import urllib.parse
import urllib.request
from pathlib import Path

from dotenv import load_dotenv
from fastmcp import FastMCP

_BASE = Path(__file__).resolve().parent
load_dotenv(_BASE / ".env")
# Use E:/ so Windows resolves to the drive root (Path("E:") alone can be wrong)
_E_FOLDER = Path("E:/")


# ---- Gmail ----
def _gmail_service():
    from google.oauth2.credentials import Credentials
    from google_auth_oauthlib.flow import InstalledAppFlow
    from google.auth.transport.requests import Request
    from googleapiclient.discovery import build

    cred_path = os.environ.get("GMAIL_CREDENTIALS_JSON", str(_BASE / "gmail_credentials.json"))
    token_path = os.environ.get("GMAIL_TOKEN_JSON", str(_BASE / "gmail_token.json"))
    scopes = ["https://www.googleapis.com/auth/gmail.readonly"]

    creds = None
    if os.path.exists(token_path):
        creds = Credentials.from_authorized_user_file(token_path, scopes)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            if not os.path.exists(cred_path):
                raise FileNotFoundError(f"Gmail credentials not found at {cred_path}")
            flow = InstalledAppFlow.from_client_secrets_file(cred_path, scopes)
            creds = flow.run_local_server(port=0)
        with open(token_path, "w") as f:
            f.write(creds.to_json())
    return build("gmail", "v1", credentials=creds)


# ---- Outlook (Microsoft Graph) ----
def _outlook_token_for_cache(cache_path: str):
    """Get (access_token, account_email) for one Outlook token cache. Raises if no token."""
    import msal

    client_id = os.environ.get("OUTLOOK_CLIENT_ID")
    if not client_id:
        raise FileNotFoundError("Set OUTLOOK_CLIENT_ID (Azure app registration, public client).")
    tenant = os.environ.get("OUTLOOK_TENANT_ID", "common")
    scopes = ["https://graph.microsoft.com/Mail.Read"]
    redirect_port = int(os.environ.get("OUTLOOK_REDIRECT_PORT", "8400"))

    cache = msal.SerializableTokenCache()
    if os.path.exists(cache_path):
        cache.deserialize(open(cache_path, "r").read())
    app = msal.PublicClientApplication(
        client_id, authority=f"https://login.microsoftonline.com/{tenant}", token_cache=cache
    )
    accounts = app.get_accounts()
    if accounts:
        result = app.acquire_token_silent(scopes, account=accounts[0])
    else:
        result = app.acquire_token_interactive(scopes, port=redirect_port)
    if not result:
        raise RuntimeError("Failed to get Outlook token.")
    open(cache_path, "w").write(cache.serialize())
    email = (accounts[0].get("username") or "") if accounts else ""
    return result["access_token"], email


def _outlook_token():
    """Single-account: returns access_token only (backward compatible)."""
    cache_path = os.environ.get("OUTLOOK_TOKEN_CACHE", str(_BASE / "outlook_token_cache.bin"))
    token, _ = _outlook_token_for_cache(cache_path)
    return token


def _outlook_all_tokens() -> list[tuple[str, str]]:
    """Returns list of (account_email, access_token) for all configured Outlook caches (main + optional second)."""
    cache1 = os.environ.get("OUTLOOK_TOKEN_CACHE", str(_BASE / "outlook_token_cache.bin"))
    cache2 = os.environ.get("OUTLOOK_TOKEN_CACHE_2")
    paths = [cache1]
    if cache2:
        paths.append(cache2)
    result: list[tuple[str, str]] = []
    last_error: Exception | None = None
    for cache_path in paths:
        try:
            token, email = _outlook_token_for_cache(cache_path)
            result.append((email or cache_path, token))
        except Exception as e:
            last_error = e
            if len(paths) == 1:
                raise
            continue
    return result


def _search_gmail_rows(
    query: str,
    max_results: int,
    after_date: str | None = None,
) -> list[dict]:
    """Same params as _search_gmail; returns list of dicts with provider, from, subject, date, snippet, link, query_used."""
    service = _gmail_service()
    q = query
    if after_date:
        q = f"{query} after:{after_date.replace('-', '/')}"
    limit = min(max_results, 100)
    results = (
        service.users()
        .messages()
        .list(userId="me", q=q, maxResults=limit)
        .execute()
    )
    rows: list[dict] = []
    for m in results.get("messages", []):
        mid = m["id"]
        try:
            msg = (
                service.users()
                .messages()
                .get(
                    userId="me",
                    id=mid,
                    format="metadata",
                    metadataHeaders=["From", "Subject", "Date"],
                )
                .execute()
            )
        except Exception:
            continue
        payload = msg.get("payload", {}) or {}
        headers = {h["name"].lower(): h["value"] for h in payload.get("headers", [])}
        snippet = (msg.get("snippet") or "").replace("\n", " ").strip()[:300]
        rows.append({
            "provider": "Gmail",
            "account": "",
            "from": headers.get("from", ""),
            "subject": headers.get("subject", ""),
            "date": headers.get("date", ""),
            "snippet": snippet,
            "link": f"https://mail.google.com/mail/u/0/#inbox/{mid}",
            "query_used": query,
        })
    return rows


def _search_outlook_rows_with_token(
    token: str,
    account_email: str,
    query: str,
    max_results: int,
    after_date: str | None = None,
) -> list[dict]:
    """Search one Outlook account (given token); returns rows with provider, account, from, subject, etc."""
    top = min(max_results * 2 if after_date else max_results, 100)
    params = [
        ("$search", f'"{query}"'),
        ("$top", str(top)),
        ("$select", "id,from,subject,receivedDateTime,bodyPreview,webLink"),
    ]
    url = "https://graph.microsoft.com/v1.0/me/messages?" + urllib.parse.urlencode(
        params, safe="$,()", quote_via=urllib.parse.quote
    )
    req = urllib.request.Request(url, headers={"Authorization": f"Bearer {token}"})
    with urllib.request.urlopen(req) as resp:
        data = json.load(resp)
    cutoff = f"{after_date}T00:00:00Z" if after_date else None
    rows: list[dict] = []
    for m in data.get("value", []):
        if len(rows) >= max_results:
            break
        received = m.get("receivedDateTime") or ""
        if cutoff and received < cutoff:
            continue
        from_obj = m.get("from", {}) or {}
        from_addr = from_obj.get("emailAddress", {}) or {}
        from_str = from_addr.get("address", "") or (from_addr.get("name", "") or "")
        mid = m.get("id", "")
        snippet = (m.get("bodyPreview") or "").replace("\n", " ").strip()[:300]
        link = m.get("webLink") or f"https://outlook.live.com/mail/0/inbox/id/{mid}"
        rows.append({
            "provider": "Outlook",
            "account": account_email,
            "from": from_str,
            "subject": m.get("subject", ""),
            "date": received,
            "snippet": snippet,
            "link": link,
            "query_used": query,
        })
    return rows


def _search_outlook_rows(
    query: str,
    max_results: int,
    after_date: str | None = None,
) -> list[dict]:
    """Search all configured Outlook accounts; returns merged rows with account column."""
    tokens = _outlook_all_tokens()
    if not tokens:
        raise RuntimeError(
            "Outlook: no valid token. One or both caches may have expired. "
            "Run the MCP again and complete browser sign-in when prompted, or run "
            "python custom_mcp_server.py and trigger an Outlook search to re-auth."
        )
    rows: list[dict] = []
    for account_email, token in tokens:
        rows.extend(
            _search_outlook_rows_with_token(token, account_email, query, max_results, after_date)
        )
    return rows


def _filter_tax_filing_only(rows: list[dict]) -> list[dict]:
    """Keep only rows that look like property tax, tax filing, or official government tax; drop payment/receipt tax."""
    filing_keywords = (
        "property tax", "tax return", "irs", "1040", "w-2", "w2", "file taxes", "tax filing",
        "tax assessment", "department of revenue", "tax form", "tax authority", "income tax",
        "form 1040", "extension", "tax deadline", "filing deadline", "federal tax", "state tax",
        "tax refund", "1099", "tax year", "property tax assessment", "assessor", "tax bill",
    )
    kept: list[dict] = []
    for r in rows:
        text = (r.get("subject", "") + " " + r.get("snippet", "") + " " + r.get("from", "")).lower()
        if any(f in text for f in filing_keywords):
            kept.append(r)
    return kept


def _write_emails_excel(rows: list[dict], path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Emails"
    headers = ["Provider", "Account", "From", "Subject", "Date", "Snippet", "Open link", "Search query used"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)
    for i, r in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=r.get("provider", ""))
        ws.cell(row=i, column=2, value=r.get("account", ""))
        ws.cell(row=i, column=3, value=r.get("from", ""))
        ws.cell(row=i, column=4, value=r.get("subject", ""))
        ws.cell(row=i, column=5, value=r.get("date", ""))
        ws.cell(row=i, column=6, value=r.get("snippet", ""))
        link = r.get("link", "")
        cell = ws.cell(row=i, column=7, value="Open in mail" if link else "")
        if link:
            cell.hyperlink = link
            cell.font = Font(color="0563C1", underline="single")
        ws.cell(row=i, column=8, value=r.get("query_used", ""))
    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = max(12, min(50, 18 if col == 6 else 25))
    ws.column_dimensions["F"].width = 50
    wb.save(path)


# ---- MCP ----
mcp = FastMCP("Email (Gmail + Outlook)")


@mcp.tool()
def search_emails_to_excel(
    provider: str,
    query: str,
    max_results: int = 50,
    after_date: str | None = None,
    output_path: str | None = None,
    tax_filing_only: bool = False,
) -> str:
    """
    Search Gmail and/or Outlook and write an Excel file with: Provider, Account, From, Subject, Date, Snippet, Open link, Search query used.
    provider: 'gmail', 'outlook', or 'both'. query: search string (e.g. 'tax return OR property tax OR IRS').
    tax_filing_only: if True, keep only emails about property tax, tax filing, or official government tax; drop payment/receipt tax.
    output_path: optional path for the .xlsx file; default is search_emails.xlsx in the server directory.
    Returns the absolute path to the saved file and the number of emails included, or an error message.
    """
    path = Path(output_path).resolve() if output_path else _BASE / "search_emails.xlsx"
    if path.suffix.lower() != ".xlsx":
        path = path.with_suffix(".xlsx")
    all_rows: list[dict] = []
    errors: list[str] = []
    p = provider.lower()
    if p in ("gmail", "both"):
        try:
            all_rows.extend(_search_gmail_rows(query, max_results, after_date))
        except FileNotFoundError as e:
            errors.append(f"Gmail: {e}")
        except Exception as e:
            errors.append(f"Gmail: {e}")
    if p in ("outlook", "both"):
        try:
            all_rows.extend(_search_outlook_rows(query, max_results, after_date))
        except FileNotFoundError as e:
            errors.append(f"Outlook: {e}")
        except Exception as e:
            errors.append(f"Outlook: {e}")
    if p not in ("gmail", "outlook", "both"):
        return f"Unknown provider: {provider}. Use 'gmail', 'outlook', or 'both'."
    before_count = 0
    if tax_filing_only and all_rows:
        before_count = len(all_rows)
        all_rows = _filter_tax_filing_only(all_rows)
    if not all_rows:
        return "No messages found for that query. " + (" ".join(errors) if errors else "")
    try:
        _write_emails_excel(all_rows, path)
        msg = f"Saved {len(all_rows)} emails to {path}"
        if tax_filing_only and before_count and before_count != len(all_rows):
            msg += f" (filtered from {before_count} to filing/official tax only)"
        return msg
    except Exception as e:
        return f"Error writing Excel: {e}"


def _read_folder_contents(folder: Path, max_chars: int = 200_000) -> str:
    """Read text/md and PDF from folder; return concatenated content."""
    lines = [f"Folder: {folder}", ""]
    text_ext = {".txt", ".md", ".markdown"}
    try:
        import pypdf
    except ImportError:
        try:
            import PyPDF2 as pypdf
        except ImportError:
            pypdf = None
    for f in sorted(folder.rglob("*")):
        if not f.is_file():
            continue
        ext = f.suffix.lower()
        if ext in text_ext:
            try:
                content = f.read_text(encoding="utf-8", errors="replace")
                lines.append(f"--- {f.name} ---")
                lines.append(content[:30_000])
                lines.append("")
            except Exception as e:
                lines.append(f"--- {f.name} (error: {e}) ---")
                lines.append("")
        elif ext == ".pdf" and pypdf:
            try:
                reader = pypdf.PdfReader(str(f))
                parts = []
                for page in reader.pages[:500]:
                    parts.append(page.extract_text() or "")
                content = "\n".join(parts)[:30_000]
                lines.append(f"--- {f.name} (PDF) ---")
                lines.append(content)
                lines.append("")
            except Exception as e:
                lines.append(f"--- {f.name} (PDF error: {e}) ---")
                lines.append("")
        elif ext == ".pdf":
            lines.append(f"--- {f.name} --- (install pypdf to extract text)")
            lines.append("")
    result = "\n".join(lines)
    if len(result) > max_chars:
        result = result[:max_chars] + "\n\n... (truncated)"
    return result


@mcp.tool()
def get_quiz_materials(school_name: str, course_name: str, quiz_id: str) -> str:
    """
    Get all quiz materials for a course: lists and reads text/md files from the quiz folder.
    school_name: e.g. 'ubc'. course_name: e.g. 'cpsc440'. quiz_id: e.g. '5' or 'quiz5'.
    Looks in E:\\academics\\{school}\\{course}\\quiz\\{quiz_id} or E:\\academics\\{school}\\{course}\\quiz{quiz_id}.
    Returns concatenated file contents for study.
    """
    base = _E_FOLDER / "academics" / school_name.lower().replace(" ", "") / course_name.lower().replace(" ", "") / "lectures"
    if not base.exists():
        return f"No course directory found: {base}"

    candidates = [
        base / f"quiz {quiz_id}",
        base / "quiz" / quiz_id,
        base / "quiz" / f"quiz{quiz_id}",
        base / f"quiz{quiz_id}",
        base / f"quiz_{quiz_id}",
        base / "lectures",
    ]
    folder = None
    for c in candidates:
        if c.exists() and c.is_dir():
            folder = c
            break
    if folder is None:
        return f"No quiz folder found. Tried: {[str(c) for c in candidates]}. Listing base: {[p.name for p in base.iterdir()]}"
    return _read_folder_contents(folder)


if __name__ == "__main__":
    mcp.run(transport="stdio")
