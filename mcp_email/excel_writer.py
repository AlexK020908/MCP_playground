"""
Write email search rows to an Excel file with headers and hyperlinks.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def write_emails_excel(rows: list[dict], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Emails"
    headers = ["Provider", "Account", "From", "Subject", "Date", "Snippet", "Open link", "Search query used"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)
    for i, r in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=r.get("provider", ""))
        ws.cell(row=i, column=2, value=r.get("account", ""))
        ws.cell(row=i, column=3, value=r.get("from", ""))
        ws.cell(row=i, column=4, value=r.get("subject", ""))
        ws.cell(row=i, column=5, value=r.get("date", ""))
        ws.cell(row=i, column=6, value=r.get("snippet", ""))
        link = r.get("link", "")
        cell = ws.cell(row=i, column=7, value="Open in mail" if link else "")
        if link:
            cell.hyperlink = link
            cell.font = Font(color="0563C1", underline="single")
        ws.cell(row=i, column=8, value=r.get("query_used", ""))
    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = max(12, min(50, 18 if col == 6 else 25))
    ws.column_dimensions["F"].width = 50
    wb.save(path)
